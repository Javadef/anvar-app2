import logging
import pandas as pd
from telegram import Update
from telegram.ext import Application, MessageHandler, CommandHandler, filters, CallbackContext
import re
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
import uuid

async def clear_excel(update: Update, context: CallbackContext):
    if os.path.exists(EXCEL_FILE):
        try:
            # Excel faylini yangidan yaratish
            df = pd.DataFrame(columns=[ "Дата", "Транспорт", "Водитель", "Телефон", "Грузоотправитель", "Грузополучатель",
        "Объем", "Kub Narxi", "Rejim", "Narx", "ID"])
            df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
            await update.message.reply_text("Excel fayli tozalandi.")
        except Exception as e:
            await update.message.reply_text(f"Xato yuz berdi: {str(e)}")
    else:
        await update.message.reply_text("Excel fayli mavjud emas.")


# Mashinalar ro'yxati
truck = {
    520: '01 Y 520 VC',
    136: '01 136 QKA', 137: '01 137 QKA', 516: '01 516 PKA', 517: '01 517 PKA',
    429: '01 429 QKA', 430: '01 430 QKA', 431: '01 431 QKA', 67: '01 067 OMA',
    324: '01 324 GMA', 325: '01 325 GMA', 643: '01 643 LKA', 645: '01 645 LKA',
    725: '01 725 LKA', 913: '01 913 OKA', 914: '01 914 OKA', 573: '01 573 LKA',
    574: '01 574 LKA', 904: '01 904 PKA',  372 : '01 372 ZLA',
    314 : '01 314 JGA',
}

# Shofyorlar ro'yxati
drivers = {
    'Шокир': 'Абдраимов Шокир Абрахимович',
    'Бурхон': 'Каримов Бурхон Кобилжон ўгли',
    'Бахтиер': 'Мираипов Бахтиёр Мирраимович',
    'Фарход': 'Акрамов Фарход Хошимджанович',
    'Тимур': 'Абдукаримов Тимур Умаралиевич',
    'Бахром': 'Караханов Бахром Бурханжонович',
    'Асатилла': 'Тулаганов Асатилла Рахматиллаевич',
    'Марат': 'Ахмедов Марат Рифкатович',
    'Рустам': 'Джалалов Рустам Мирзатиллаевич',
    'Шокир': 'Лапасов Шокир Абраевич',
    'Мирфозил': 'Таджиев Мирфозил Миртоирович',
    'Дилшод': 'Умирзаков Дилшод Касимжанович',
    'Файзулло': 'Каимов Файзулла Хамидуллаевич',
    'Искандар': 'Агзамов Искандар Аскарджанович',
    'Хусан': 'Кодиров Хусан',
    'Аваз' : 'Бабаев Аваз Рихсибаевич',
    'Мехриддин' : 'Наджимов Мехриддин',
    'Ихтиер' : 'Рахманалийев Ихтиёор',
    'Умид': 'Умирзаков Умид Адилович',
}

region_prices = {
    'Жиззах': {
        '8': 2520100, '18':  2897950, '34':  3630000, '42':  3850000, '47':  5376800, '50':  5720000, '56':  7166992, '60':  7678920, '88':  8446850, },
    'Термез': {
        '8': 5311900, '18': 6109400, '34': 7150000, '42': 7590000, '47': 11580800, '50': 12320000, '56': 16336572, '60': 17503470, '88': 19253830, },
    'Самарканд': {
        '8':  2825900, '18':  3250500, '34': 7150000, '42':  4290000, '47':  5790400, '50':  6160000, '56':  7904848, '60':  8469480, '88':  9316380, },
    'Бухоро': {
    '8': 441501, '18': 4738800, '34': 5610000,'42': 654107, '47': 7858400, '50': 8360000,'56': 1129256, '60': 11292560, '88': 12421830,
    
    
}
}

# Excel fayl nomi
EXCEL_FILE = 'zayavka.xlsx'

# Telefon raqamini chiqarish funksiyasi
def getPhoneNum(message):
    phone_match = re.search(r"тел[:\s]*([\+\d\(\)\-\s]+)", message)
    if phone_match:
        phone_number = phone_match.group(1)
        cleaned_phone_number = re.sub(r"[^\d+]", "", phone_number)
        return cleaned_phone_number
    return None

# Xaydovchi ismini olish
def getDriver(driver_name):
    driver_name = driver_name.strip()
    if driver_name.startswith('Вод.') or driver_name.startswith('Вод'):
        driver_name = driver_name[4:].strip()
    driver_name = re.sub(r"тел[:\s]*([\+\d\(\)\-\s]+)", "", driver_name).strip()
    for key in drivers:
        if key.lower() in driver_name.lower():
            return drivers[key]
    return "Shofyor topilmadi"

# Mashina raqamini olish
def getCarNum(message):
    truck_match = re.search(r"Вид транспорта\s*([A-Za-zа-яА-Я0-9\s]+)", message)
    if truck_match:
        truck_number = truck_match.group(1).strip()
        truck_number_match = re.search(r"(\d{3})", truck_number)
        if truck_number_match:
            truck_id = int(truck_number_match.group(1))
            return truck.get(truck_id, "Mashina nomi topilmadi")
    return "Mashina raqami topilmadi"

# Sana olish
def extractDate(message):
    date_match = re.search(r"Дата\s*([\d\.]+)", message)
    if date_match:
        return date_match.group(1).strip()
    return "Sana topilmadi"

# Zayavkani parsing qilish funksiyasi
# Zayavkani parsing qilish funksiyasi
# Zayavka parsing qilish funksiyasi
# Zayavka parsing qilish funksiyasi
# Foydalanuvchi javobini aniqlash va ko'rsatkichlarni kuzatish
async def parse_message(message, update: Update, context: CallbackContext):
    data = {}
    data['Дата'] = extractDate(message)
    data['Водитель'] = getDriver(message)
    data['Транспорт'] = getCarNum(message)
    data['Телефон'] = getPhoneNum(message)
    region_match = re.search(r"Грузополучатель\s*(.*)", message)
    data['Грузополучатель'] = region_match.group(1).strip() if region_match else "Ma'lumot yo'q"
    data['Объем'] = extractVolume(message)  # So‘rovdagi hajm
    data['ID'] = str(uuid.uuid4())[:8]

    region = data['Грузополучатель']
    context.user_data['data'] = data  # Vaqtinchalik ma'lumotlar

    if region in region_prices:
        # Kub narxini so'rash
        await update.message.reply_text("Куб нархини киритинг:")
        context.user_data['awaiting_kub'] = True
    else:
        data['Kub Narxi'] = "Narx topilmadi"
        data['Narx'] = "Narx topilmadi"
        save_to_excel(data)
        await update.message.reply_text("Narx topilmadi, ma'lumot saqlandi.")
# Excelga saqlash funksiyasi

# Zayavkani qabul qilish va saqlash
# Zayavkani qabul qilish va saqlash
# Zayavka qabul qilish va saqlash
async def handle_message(update: Update, context: CallbackContext):
    message = update.message.text
    try:
        # Kub narxi uchun foydalanuvchi javobi
        if context.user_data.get('awaiting_kub'):
            try:
                user_kub_price = message  # Foydalanuvchi yuborgan qiymat
                data = context.user_data['data']

                # Excel uchun ma'lumotni to'ldirish
                data['Kub Narxi'] = user_kub_price
                save_to_excel(data)  # Excelga yozish

                # Javob yuborish
                response_text = "\n".join([f"{key}: {value}" for key, value in data.items()])
                await update.message.reply_text(f"Zayavka muvaffaqiyatli saqlandi:\n{response_text}")

                # Bayroqni o'chirish
                context.user_data['awaiting_kub'] = False
                context.user_data['data'] = None
            except ValueError:
                await update.message.reply_text("Iltimos, to‘g‘ri narx.")

        # Zayavka uchun birinchi bosqich
        else:
            await parse_message(message, update, context)
    except Exception as e:
        await update.message.reply_text(f"Xato yuz berdi: {str(e)}")


async def parse_message(message, update: Update, context: CallbackContext):
    data = {}
    data['Дата'] = extractDate(message)
    data['Водитель'] = getDriver(message)
    data['Транспорт'] = getCarNum(message)
    data['Телефон'] = getPhoneNum(message)
    data['Грузоотправитель'] = 'Тошкент'
    # Грузополучатель va Объем ma'lumotlari
    region_match = re.search(r"Грузополучатель\s*(.*)", message)
    data['Грузополучатель'] = region_match.group(1).strip() if region_match else "Ma'lumot yo'q"
    volume_match = re.search(r"Объем груза\s*([\d\.]+)", message)
    
    if volume_match:
        try:
            raw_volume = volume_match.group(1).strip()
            data['Объем'] = clean_volume_input(raw_volume)  # Tozalangan qiymatni yozish
        except ValueError:
            data['Объем'] = 0  # Agar qiymat noto'g'ri bo'lsa, 0 sifatida belgilash
    else:
        data['Объем'] = 0  # Agar qiymat topilmasa

    data['ID'] = str(uuid.uuid4())[:8]

    # Ma'lumotlarni vaqtinchalik saqlash va foydalanuvchidan Kub Narxi so‘rash
    context.user_data['data'] = data
    context.user_data['awaiting_kub'] = True
    await update.message.reply_text("Kub narxini kiriting:")

def extractVolume(message):
    volume_match = re.search(r"Объем груза\s*([\d\.]+)", message)
    if volume_match:
        return float(volume_match.group(1).strip())
    return "Ma'lumot yo'q"

def get_price(region, volume):
    # Region va hajmga mos narxni olish
    volume_int = round(volume)
    price =  region_prices.get(region, {}).get(volume_int)
    return price



# Excel faylni yuborish
async def send_file(update: Update, context: CallbackContext):
    if os.path.exists(EXCEL_FILE):
        await update.message.reply_document(document=open(EXCEL_FILE, 'rb'))
    else:
        await update.message.reply_text(f"{EXCEL_FILE} fayli mavjud emas.")

# Kub hajmini olish va saqlash
# Kub hajmini olish va saqlash
async def get_kub(update: Update, context: CallbackContext):
    try:
        # Kub miqdorini foydalanuvchidan olish
        user_kub = float(update.message.text)
        data = context.user_data.get('data')

        if data:
            region = data['Грузополучатель']
            region_data = region_prices.get(region, {})

            # Kub hajmini va narxni hisoblash
            if region_data:
                closest_kub = min(region_data.keys(), key=lambda x: abs(x - user_kub))
                data['Объем'] = user_kub
                data['Kub Narxi'] = closest_kub
                data['Narx'] = region_data[closest_kub]
            else:
                data['Объем'] = user_kub
                data['Kub Narxi'] = "Narx topilmadi"
                data['Narx'] = "Narx topilmadi"

            save_to_excel(data)  # Excelga yozish

            # Ma'lumotni qaytarish
            response_text = "\n".join([f"{key}: {value}" for key, value in data.items()])
            await update.message.reply_text(f"Zayavka muvaffaqiyatli saqlandi:\n{response_text}")

            # Kubni olishni yakunlash
            context.user_data['data'] = None  # Kub so'rovini yakunlash

        else:
            await update.message.reply_text("Ma'lumotni olishda xatolik yuz berdi.")
    except ValueError:
        # Agar foydalanuvchi raqamdan boshqa narsa kiritsa
        await update.message.reply_text("Iltimos, to'g'ri kub hajmini kiriting.")
    except Exception as e:
        await update.message.reply_text(f"Xato yuz berdi: {str(e)}")

def clean_volume_input(value):
    value = value.strip()  # Bo'sh joylarni olib tashlash
    if value.startswith('.'):  # Boshida nuqta bo'lsa olib tashlash
        value = value[1:]
    if value.endswith('.'):  # Oxirida nuqta bo'lsa olib tashlash
        value = value[:-1]
    return float(value)  # Tozalangan qiymatni float formatiga aylantirish


def save_to_excel(data):

    zayavka_region = data.get("Грузополучатель", "")
    zayavka_kub = data.get("Kub Narxi", "")
    if zayavka_region in region_prices:
        print(f"{zayavka_region} regioni topildi.")
        print(f"Regiondagi kub hajmlari: {region_prices[zayavka_region].keys()}")  # Regiondagi mavjud kub hajmlarini chiqarish
    
    if zayavka_kub in region_prices[zayavka_region]:
        narx = region_prices[zayavka_region][zayavka_kub]
        print(f"Region: {zayavka_region}, Kub: {zayavka_kub}, Narx: {narx}")
    else: narx = 0

    """
    Ma'lumotlarni Excel fayliga saqlash funksiyasi.
    """
    # Yangi tartibga ko'ra ustunlarni belgilash
    columns = [
        "Дата", "Транспорт", "Водитель", "Телефон", "Грузоотправитель", "Грузополучатель",
        "Объем", "Kub Narxi", "Rejim", "Narx", "ID"
    ]
    
    # Excel fayli mavjudligini tekshirish
    if os.path.exists(EXCEL_FILE):
        try:
            df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
        except Exception as e:
            print(f"Excel faylni o'qishda xato: {str(e)}")
            df = pd.DataFrame(columns=columns)  # Agar xato bo'lsa yangi fayl yaratamiz
    else:
        df = pd.DataFrame(columns=columns)  # Fayl bo'lmasa yangi fayl yaratamiz
    
    # Zayavkadan kelgan ma'lumotlarni to'g'ri joylashtirish
    # Zayavkadan kelgan ma'lumotlarni to'g'ri joylashtirish
    new_row = {
    "Дата": data.get("Дата", ""),
    "Транспорт": data.get("Транспорт", ""),
    "Водитель": data.get("Водитель", ""),
    "Телефон": data.get("Телефон", ""),
    "Грузоотправитель": "Тошкент", 
    "Грузополучатель": data.get("Грузополучатель", ""),  # "Грузополучатель"ni orqaga o'zgartirdim
    "Объем": data.get("Объем", ""),  # Zayavkadan kelgan hajm
    "Kub Narxi": data.get("Kub Narxi", ""),  # Foydalanuvchi yozgan hajm
    "Ref": "Реф",  # Statik yozib qo'yamiz
    "Narx": narx,
    "ID": data.get("ID", ""),  # UUID
}


    # Excelga yozish
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    
    try:
        # Excel fayliga yozish
        df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
    except Exception as e:
        print(f"Excelga yozishda xato: {str(e)}")
    
    try:
        # Excelda chizgilar qo'shish
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = Border(
                    left=Side(border_style="thin"), 
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"), 
                    bottom=Side(border_style="thin")
                )
        workbook.save(EXCEL_FILE)
    except Exception as e:
        print(f"Chizgilarni qo'shishda xato: {str(e)}")


# Handlerlarni qo'shish
def main():
    API_TOKEN = '7212744246:AAGKI7Qp8fncrz66PuI2OZQHEmEMncAt7Ww'
    application = Application.builder().token(API_TOKEN).build()

    # Handlerlarni qo'shish
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_handler(CommandHandler("clear", clear_excel))
    application.add_handler(CommandHandler("sendF", send_file))

    application.run_polling()

if __name__ == '__main__':
    logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
    main()
